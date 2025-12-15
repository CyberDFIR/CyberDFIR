

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font


def report(stock_data):
    if not stock_data:
        print("No data to report.")
        return

    df = pd.DataFrame(stock_data)

    df["Change %"] = (
        (df["Current Price"] - df["Previous Close"])
        / df["Previous Close"]
    ) * 100

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"Market_Report_{today}.xlsx"
    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in ws.iter_rows(min_row=2):
        change = row[3].value

        if change > 0:
            font = Font(color="008000")
        elif change < 0:
            font = Font(color="FF0000")
        else:
            continue

        for cell in row:
            cell.font = font

    wb.save(filename)
    print(f"Report created: {filename}")
